# -*- coding: utf-8 -*-
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .project_parser import parse_date


HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
CGI_RED_FILL = PatternFill("solid", fgColor="E31937")
ACTIVE_FILL = PatternFill("solid", fgColor="FADCE1")
DONE_FILL = PatternFill("solid", fgColor="D9F2E3")
SOON_FILL = PatternFill("solid", fgColor="FFF1CC")
OVERDUE_FILL = PatternFill("solid", fgColor="F8D7DA")
THIN = Side(style="thin", color="B7B7B7")
ALL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
EXPORT_VIEW_CONFIG = {
    "active": {
        "sheet_name": "Acoes em Progresso",
        "title": "O que esta a acontecer agora",
        "subtitle": "Acoes em curso ordenadas pelos prazos mais proximos",
        "label": "Em progresso",
    },
    "risk": {
        "sheet_name": "Acoes em Risco",
        "title": "O que esta em risco",
        "subtitle": "Items com atraso ou prazos muito proximos",
        "label": "Em risco",
    },
    "upcoming": {
        "sheet_name": "Proximas Acoes",
        "title": "O que vai acontecer a seguir",
        "subtitle": "Acoes previstas para iniciar de seguida",
        "label": "Proxima vaga",
    },
}


def mondays_between(start: date, end: date) -> list[date]:
    current = start - timedelta(days=start.weekday())
    weeks: list[date] = []
    while current <= end:
        weeks.append(current)
        current += timedelta(days=7)
    return weeks


def task_fill(task: dict) -> PatternFill:
    bucket = task.get("bucket")
    if bucket == "completed":
        return DONE_FILL
    if bucket == "overdue":
        return OVERDUE_FILL
    if bucket == "due_soon":
        return SOON_FILL
    return ACTIVE_FILL


def style_cell(cell, fill=None, bold=False, size=11):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = ALL_BORDER
    cell.font = Font(name="Calibri", size=size, bold=bold)
    if fill:
        cell.fill = fill


def unique_sheet_name(workbook: Workbook, raw_name: str) -> str:
    base = (raw_name or "Release").strip()[:31] or "Release"
    candidate = base
    counter = 2
    while candidate in workbook.sheetnames:
        suffix = f" {counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    return candidate


def sort_tasks_for_export(tasks: list[dict], mode: str) -> list[dict]:
    if mode == "active":
        return sorted(
            [task for task in tasks if task.get("bucket") == "active"],
            key=lambda task: (task.get("end") or "9999-12-31", task.get("start") or "9999-12-31", task.get("key") or ""),
        )
    if mode == "risk":
        priority = {"overdue": 0, "due_soon": 1}
        return sorted(
            [task for task in tasks if task.get("bucket") in {"overdue", "due_soon"}],
            key=lambda task: (priority.get(task.get("bucket"), 9), task.get("end") or "9999-12-31", task.get("key") or ""),
        )
    if mode == "upcoming":
        return sorted(
            [task for task in tasks if task.get("bucket") == "upcoming"],
            key=lambda task: (task.get("start") or "9999-12-31", task.get("end") or "9999-12-31", task.get("key") or ""),
        )
    return list(tasks)


def write_sheet_title(ws, row: int, title: str, subtitle: str | None = None) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    title_cell = ws.cell(row, 1, title)
    style_cell(title_cell, fill=CGI_RED_FILL, bold=True, size=13)
    title_cell.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    if subtitle:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=9)
        subtitle_cell = ws.cell(row + 1, 1, subtitle)
        style_cell(subtitle_cell, fill=HEADER_FILL, size=10)
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        return row + 3
    return row + 2


def write_summary_metric(ws, row: int, col: int, label: str, value: str | int | float) -> None:
    label_cell = ws.cell(row, col, label)
    style_cell(label_cell, fill=HEADER_FILL, bold=True, size=10)
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    value_cell = ws.cell(row, col + 1, value)
    style_cell(value_cell, bold=True, size=12)
    value_cell.alignment = Alignment(horizontal="left", vertical="center")


def populate_task_sheet(ws, title: str, subtitle: str, tasks: list[dict], view_label: str) -> None:
    row = write_sheet_title(ws, 1, title, subtitle)
    headers = ["Key", "Title", "Parent", "Assignee", "Status", "Leitura", "Start", "End", "Release", "Type", "Notes"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row, col, header)
        style_cell(cell, fill=HEADER_FILL, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    if not tasks:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=len(headers))
        info_cell = ws.cell(row + 1, 1, "Sem items para esta vista.")
        style_cell(info_cell)
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        for row_idx, task in enumerate(tasks, start=row + 1):
            values = [
                task.get("key"),
                task.get("title"),
                task.get("parent"),
                task.get("assignee"),
                task.get("status"),
                view_label,
                task.get("start"),
                task.get("end"),
                task.get("release"),
                task.get("type"),
                task.get("notes"),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row_idx, col_idx, value)
                style_cell(cell, fill=task_fill(task) if col_idx in (1, 5, 6) else None)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=col_idx == len(headers))

    widths = [14, 54, 42, 24, 18, 18, 12, 12, 34, 16, 56]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = f"A{row + 1}"
    ws.sheet_view.showGridLines = False


def export_project(project: dict, target_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Gantt (Weeks)"

    tasks = project["tasks"]
    metrics = project.get("metrics", {}) or {}
    active_tasks = sort_tasks_for_export(tasks, "active")
    risk_tasks = sort_tasks_for_export(tasks, "risk")
    upcoming_tasks = sort_tasks_for_export(tasks, "upcoming")
    dated_tasks = [task for task in tasks if task.get("start") and task.get("end")]
    if dated_tasks:
        min_start = min(parse_date(task["start"]) for task in dated_tasks if parse_date(task["start"]))
        max_end = max(parse_date(task["end"]) for task in dated_tasks if parse_date(task["end"]))
    else:
        today = date.today()
        min_start = today
        max_end = today + timedelta(days=35)

    weeks = mondays_between(min_start, max_end)
    static_headers = ["Key", "Title", "Type", "Parent", "Assignee", "Status", "Start", "End", "Duration (d)", "Notes"]
    first_week_col = len(static_headers) + 1

    for col, header in enumerate(static_headers, start=1):
        cell = ws.cell(1, col, header)
        style_cell(cell, fill=HEADER_FILL, bold=True)
        if col <= len(static_headers):
            ws.merge_cells(start_row=1, start_column=col, end_row=3, end_column=col)

    month_ranges: list[tuple[int, int, str]] = []
    current_month = None
    month_start_col = first_week_col
    for idx, week in enumerate(weeks, start=first_week_col):
        month_label = week.strftime("%b-%Y")
        if current_month != month_label:
            if current_month is not None:
                month_ranges.append((month_start_col, idx - 1, current_month))
            current_month = month_label
            month_start_col = idx
        ws.cell(2, idx, week.strftime("%b"))
        style_cell(ws.cell(2, idx), fill=HEADER_FILL, bold=True, size=9)
        ws.cell(3, idx, week.strftime("%d-%b-%y"))
        style_cell(ws.cell(3, idx), fill=HEADER_FILL, bold=True, size=9)
    if weeks:
        month_ranges.append((month_start_col, (first_week_col - 1) + len(weeks), current_month or weeks[0].strftime("%b-%Y")))
        for start_col, end_col, label in month_ranges:
            ws.cell(1, start_col, label)
            style_cell(ws.cell(1, start_col), fill=HEADER_FILL, bold=True)
            if end_col > start_col:
                ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    widths = [13, 64, 14, 40, 22, 18, 12, 12, 12, 52]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for col in range(first_week_col, first_week_col + len(weeks)):
        ws.column_dimensions[get_column_letter(col)].width = 4.2

    for row_idx, task in enumerate(tasks, start=4):
        values = [
            task.get("key"),
            task.get("title"),
            task.get("type"),
            task.get("parent"),
            task.get("assignee"),
            task.get("status"),
            task.get("start"),
            task.get("end"),
            task.get("duration_days"),
            task.get("notes"),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            style_cell(cell)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in (1, 2, 3, 4, 5, 6, 10) else "center",
                vertical="center",
                wrap_text=col_idx == 10,
            )
        start = parse_date(task.get("start"))
        end = parse_date(task.get("end"))
        if start and end:
            for week_idx, week in enumerate(weeks, start=first_week_col):
                week_end = week + timedelta(days=6)
                if week <= end and week_end >= start:
                    cell = ws.cell(row_idx, week_idx, "")
                    style_cell(cell, fill=task_fill(task))
                else:
                    style_cell(ws.cell(row_idx, week_idx))

    ws.freeze_panes = f"{get_column_letter(first_week_col)}4"
    ws.sheet_view.showGridLines = False

    summary = wb.create_sheet("Executive Summary", 1)
    next_row = write_sheet_title(
        summary,
        1,
        "Resumo Executivo",
        f"{project.get('project', 'Projeto')} | atualizado {project.get('updated_at', project.get('imported_at', 'n/d'))}",
    )
    write_summary_metric(summary, next_row, 1, "Projeto", project.get("project", "n/d"))
    write_summary_metric(summary, next_row + 1, 1, "Origem", str(project.get("source_type", "n/d")).upper())
    write_summary_metric(summary, next_row + 2, 1, "Total de items", metrics.get("total", len(tasks)))
    write_summary_metric(summary, next_row, 4, "Progresso global", f"{metrics.get('weighted_completion_ratio', 0)}%")
    write_summary_metric(summary, next_row + 1, 4, "Progresso releases", f"{metrics.get('weighted_release_completion_ratio', metrics.get('release_completion_ratio', 0))}%")
    write_summary_metric(summary, next_row + 2, 4, "Peso releases dev", f"{project.get('settings', {}).get('release_weight', metrics.get('release_weight', 0))}%")
    write_summary_metric(summary, next_row, 7, "Em curso agora", len(active_tasks))
    write_summary_metric(summary, next_row + 1, 7, "Em risco", len(risk_tasks))
    write_summary_metric(summary, next_row + 2, 7, "A seguir", len(upcoming_tasks))

    section_row = next_row + 5
    section_row = write_sheet_title(summary, section_row, "Vistas exportadas", "Blocos operacionais incluidos automaticamente no Excel")
    export_headers = ["Vista", "Total", "Primeiro item", "Deadline / inicio"]
    for col, header in enumerate(export_headers, start=1):
        cell = summary.cell(section_row, col, header)
        style_cell(cell, fill=HEADER_FILL, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    export_rows = [
        ("Em progresso", active_tasks),
        ("Em risco", risk_tasks),
        ("Proximas acoes", upcoming_tasks),
    ]
    for row_idx, (label, view_tasks) in enumerate(export_rows, start=section_row + 1):
        first_task = view_tasks[0] if view_tasks else {}
        values = [
            label,
            len(view_tasks),
            f"{first_task.get('key', '')} - {first_task.get('title', '')}".strip(" -"),
            first_task.get("end") or first_task.get("start") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = summary.cell(row_idx, col_idx, value or "n/d")
            style_cell(cell, fill=task_fill(first_task) if col_idx in (1, 2) and first_task else None)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    section_row += len(export_rows) + 4
    section_row = write_sheet_title(summary, section_row, "Saude das Releases", "Leitura executiva das releases de desenvolvimento")
    release_headers = ["Release", "Efetivo %", "Desenv. %", "Suporte %", "Peso bloco %", "Items"]
    for col, header in enumerate(release_headers, start=1):
        cell = summary.cell(section_row, col, header)
        style_cell(cell, fill=HEADER_FILL, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    release_breakdown = metrics.get("release_breakdown", []) or []
    if not release_breakdown:
        summary.merge_cells(start_row=section_row + 1, start_column=1, end_row=section_row + 1, end_column=len(release_headers))
        info_cell = summary.cell(section_row + 1, 1, "Sem releases de desenvolvimento identificadas.")
        style_cell(info_cell)
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
        section_row += 3
    else:
        for row_idx, release in enumerate(release_breakdown, start=section_row + 1):
            values = [
                release.get("name"),
                f"{release.get('effective_completion_ratio', release.get('completion_ratio', 0))}%",
                f"{release.get('development_completion_ratio', 0)}%",
                f"{release.get('support_completion_ratio', 0)}%",
                f"{release.get('normalized_weight', 0)}%",
                release.get("total", 0),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = summary.cell(row_idx, col_idx, value)
                style_cell(cell)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        section_row += len(release_breakdown) + 3

    section_row = write_sheet_title(summary, section_row, "Responsaveis mais expostos", "Carga atual do plano")
    owner_headers = ["Responsavel", "Items no plano"]
    for col, header in enumerate(owner_headers, start=1):
        cell = summary.cell(section_row, col, header)
        style_cell(cell, fill=HEADER_FILL, bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    top_assignees = metrics.get("top_assignees", []) or []
    if not top_assignees:
        summary.merge_cells(start_row=section_row + 1, start_column=1, end_row=section_row + 1, end_column=2)
        info_cell = summary.cell(section_row + 1, 1, "Sem responsaveis atribuidos para destacar.")
        style_cell(info_cell)
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        for row_idx, (assignee, count) in enumerate(top_assignees, start=section_row + 1):
            cell_name = summary.cell(row_idx, 1, assignee)
            style_cell(cell_name)
            cell_name.alignment = Alignment(horizontal="left", vertical="center")
            cell_count = summary.cell(row_idx, 2, count)
            style_cell(cell_count)
            cell_count.alignment = Alignment(horizontal="left", vertical="center")

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 28
    summary.column_dimensions["C"].width = 12
    summary.column_dimensions["D"].width = 22
    summary.column_dimensions["E"].width = 18
    summary.column_dimensions["F"].width = 12
    summary.column_dimensions["G"].width = 18
    summary.column_dimensions["H"].width = 18
    summary.column_dimensions["I"].width = 16
    summary.sheet_view.showGridLines = False

    populate_task_sheet(
        wb.create_sheet(EXPORT_VIEW_CONFIG["active"]["sheet_name"], 2),
        EXPORT_VIEW_CONFIG["active"]["title"],
        EXPORT_VIEW_CONFIG["active"]["subtitle"],
        active_tasks[:50],
        EXPORT_VIEW_CONFIG["active"]["label"],
    )
    populate_task_sheet(
        wb.create_sheet(EXPORT_VIEW_CONFIG["risk"]["sheet_name"], 3),
        EXPORT_VIEW_CONFIG["risk"]["title"],
        EXPORT_VIEW_CONFIG["risk"]["subtitle"],
        risk_tasks[:50],
        EXPORT_VIEW_CONFIG["risk"]["label"],
    )
    populate_task_sheet(
        wb.create_sheet(EXPORT_VIEW_CONFIG["upcoming"]["sheet_name"], 4),
        EXPORT_VIEW_CONFIG["upcoming"]["title"],
        EXPORT_VIEW_CONFIG["upcoming"]["subtitle"],
        upcoming_tasks[:50],
        EXPORT_VIEW_CONFIG["upcoming"]["label"],
    )

    sprint = wb.create_sheet("Sprint Allocation", 5)
    sprint["A1"] = "Sprint allocation (generated from normalized task dates)"
    style_cell(sprint["A1"], fill=CGI_RED_FILL, bold=True)
    sprint["A1"].font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    monthly = defaultdict(int)
    for task in tasks:
        end = parse_date(task.get("end"))
        if end:
            monthly[end.strftime("%b-%Y")] += 1
    sprint.append(["Period", "Tasks"])
    for cell in sprint[2]:
        style_cell(cell, fill=HEADER_FILL, bold=True)
    for period, count in monthly.items():
        sprint.append([period, count])
    for row in sprint.iter_rows(min_row=3, max_col=2):
        for cell in row:
            style_cell(cell)
    sprint.column_dimensions["A"].width = 18
    sprint.column_dimensions["B"].width = 12

    release_groups = project.get("releases") or {}
    if not release_groups:
        fallback_groups = defaultdict(list)
        for task in tasks:
            fallback_groups[task.get("release") or "Backlog"].append({"key": task.get("key"), "title": task.get("title")})
        release_groups = dict(fallback_groups)

    for release_name, release_tasks in list(release_groups.items())[:8]:
        sheet = wb.create_sheet(unique_sheet_name(wb, release_name or "Release"))
        sheet.append(["Key", "Summary"])
        for cell in sheet[1]:
            style_cell(cell, fill=HEADER_FILL, bold=True)
        for task in release_tasks:
            sheet.append([task.get("key"), task.get("title")])
        for row in sheet.iter_rows(min_row=2, max_col=2):
            for cell in row:
                style_cell(cell)
                cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.column_dimensions["A"].width = 16
        sheet.column_dimensions["B"].width = 96

    target_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target_path)
    return target_path
