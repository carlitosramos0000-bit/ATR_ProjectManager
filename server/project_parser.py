# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import io
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DATE_PATTERNS = [
    "%Y-%m-%d",
    "%d-%b-%Y",
    "%d-%b-%y",
    "%d/%m/%Y",
    "%m/%d/%Y",
    "%Y/%m/%d",
]

DONE_STATUSES = {"closed", "done", "resolved", "completed"}
ACTIVE_STATUSES = {"open", "in progress", "external pending"}
UPCOMING_STATUSES = {"ready for work", "to do", "backlog", "planned"}
DEFAULT_PROJECT_SETTINGS = {"release_weight": 90, "release_weights": {}}
RELEASE_KEYWORD_PATTERN = re.compile(r"\brel(?:e)?ase\b", re.IGNORECASE)
RELEASE_LABEL_PATTERN = re.compile(r"\brel(?:e)?ase\b(?:\s*[-:/]?\s*[\w.,]+){0,1}", re.IGNORECASE)
RELEASE_SUPPORT_PATTERN = re.compile(
    r"\b(?:test(?:e|es|ing)?|qa|uat|bug(?:\s|-)?fix(?:ing)?|re[\s-]?(?:test(?:ing)?|teste))\b",
    re.IGNORECASE,
)
RELEASE_SUPPORT_CAP = 10.0
MOJIBAKE_PATTERN = re.compile(r"(?:Ã.|Â.|â.|ðŸ)")


@dataclass
class NormalizedProject:
    project: str
    source_name: str
    imported_at: str
    updated_at: str
    source_type: str
    settings: dict[str, Any]
    tasks: list[dict[str, Any]]
    releases: dict[str, list[dict[str, Any]]]
    metrics: dict[str, Any]
    slug: str | None = None


def parse_date(value: Any) -> date | None:
    if value in (None, "", "nan"):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in DATE_PATTERNS:
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def iso_date(value: date | None) -> str | None:
    return value.isoformat() if value else None


def repair_text(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    repaired = text
    for _ in range(2):
        if not MOJIBAKE_PATTERN.search(repaired):
            break
        for encoding in ("cp1252", "latin-1"):
            try:
                candidate = repaired.encode(encoding).decode("utf-8")
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
            if candidate and candidate != repaired:
                repaired = candidate.strip()
                break
        else:
            break
    return repaired


def normalize_status(status: Any) -> str:
    if not status:
        return "Unknown"
    return repair_text(status)


def contains_release_keyword(*parts: Any) -> bool:
    return any(RELEASE_KEYWORD_PATTERN.search(str(part or "")) for part in parts)


def extract_release_label(*parts: Any) -> str:
    for part in parts:
        text = str(part or "").strip()
        if not text:
            continue
        match = RELEASE_LABEL_PATTERN.search(text)
        if match:
            label = re.sub(r"\s+", " ", match.group(0)).strip(" -:;/,")
            if label:
                return label
    return ""


def infer_release_name(task: dict[str, Any]) -> str:
    explicit_release = str(task.get("release") or "").strip()
    if explicit_release:
        return explicit_release

    parent = str(task.get("parent") or "").strip()
    title = str(task.get("title") or "").strip()
    issue_type = str(task.get("type") or "").strip()

    if contains_release_keyword(parent):
        return parent

    extracted = extract_release_label(title, issue_type)
    if extracted:
        return extracted

    if contains_release_keyword(title, issue_type):
        return "Issues de Release"

    return ""


def derive_phase(task: dict[str, Any]) -> str:
    parent = (task.get("parent") or "").strip()
    if parent:
        return parent
    release = (task.get("release") or "").strip()
    if release:
        return release
    return "Sem agrupamento"


def classify_release_component(task: dict[str, Any]) -> str:
    if not task.get("is_release_item"):
        return ""

    text = " ".join(
        [
            str(task.get("title") or "").strip(),
            str(task.get("parent") or "").strip(),
            str(task.get("type") or "").strip(),
        ]
    )
    if RELEASE_SUPPORT_PATTERN.search(text):
        return "support"
    return "development"


def classify_bucket(task: dict[str, Any], today: date) -> str:
    status = task["status"].lower()
    end = parse_date(task.get("end"))
    start = parse_date(task.get("start"))
    if status in DONE_STATUSES:
        return "completed"
    if end and end < today:
        return "overdue"
    if end and end <= today + timedelta(days=7):
        return "due_soon"
    if start and start > today:
        return "upcoming"
    if status in UPCOMING_STATUSES:
        return "upcoming"
    return "active"


def normalize_project_settings(settings: dict[str, Any] | None) -> dict[str, Any]:
    return normalize_project_settings_for_releases(settings)


def release_names_from_sources(tasks: list[dict[str, Any]], releases: dict[str, list[dict[str, Any]]] | None = None) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()

    for release_name in (releases or {}).keys():
        normalized_name = str(release_name or "").strip()
        if normalized_name and normalized_name not in seen:
            seen.add(normalized_name)
            names.append(normalized_name)

    for task in tasks:
        if not task.get("is_release_item"):
            continue
        release_name = str(task.get("release") or "Issues de Release").strip() or "Issues de Release"
        if release_name not in seen:
            seen.add(release_name)
            names.append(release_name)

    return names


def normalize_project_settings_for_releases(
    settings: dict[str, Any] | None,
    release_names: list[str] | None = None,
) -> dict[str, Any]:
    normalized = dict(DEFAULT_PROJECT_SETTINGS)
    normalized["release_weights"] = {}
    if not isinstance(settings, dict):
        if release_names:
            even_weight = round(100 / len(release_names), 2)
            normalized["release_weights"] = {name: even_weight for name in release_names}
        return normalized

    release_weight = settings.get("release_weight", DEFAULT_PROJECT_SETTINGS["release_weight"])
    try:
        release_weight = float(release_weight)
    except (TypeError, ValueError):
        release_weight = DEFAULT_PROJECT_SETTINGS["release_weight"]
    normalized["release_weight"] = max(0, min(100, round(release_weight, 1)))

    raw_release_weights = settings.get("release_weights", {})
    parsed_release_weights: dict[str, float] = {}
    if isinstance(raw_release_weights, dict):
        for raw_name, raw_value in raw_release_weights.items():
            release_name = repair_text(raw_name)
            if not release_name:
                continue
            try:
                weight_value = float(raw_value)
            except (TypeError, ValueError):
                continue
            parsed_release_weights[release_name] = max(0, round(weight_value, 2))

    normalized_release_names = [repair_text(name) for name in (release_names or []) if repair_text(name)]
    if normalized_release_names:
        positive_weights = [value for value in parsed_release_weights.values() if value > 0]
        default_weight = round(100 / len(normalized_release_names), 2)
        if positive_weights:
            default_weight = round(sum(positive_weights) / len(positive_weights), 2)
        normalized["release_weights"] = {
            release_name: parsed_release_weights.get(release_name, default_weight)
            for release_name in normalized_release_names
        }
    else:
        normalized["release_weights"] = parsed_release_weights
    return normalized


def compute_metrics(
    tasks: list[dict[str, Any]],
    releases: dict[str, list[dict[str, Any]]] | None = None,
    settings: dict[str, Any] | None = None,
) -> dict[str, Any]:
    today = date.today()
    counters = Counter()
    release_counts = Counter()
    assignee_counts = Counter()
    release_progress: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "total": 0,
            "completed": 0,
            "development_total": 0,
            "development_completed": 0,
            "support_total": 0,
            "support_completed": 0,
        }
    )
    release_names = release_names_from_sources(tasks, releases)
    normalized_settings = normalize_project_settings_for_releases(settings, release_names)
    release_tasks = 0
    release_completed = 0
    non_release_tasks = 0
    non_release_completed = 0

    for task in tasks:
        bucket = classify_bucket(task, today)
        counters[bucket] += 1
        if task.get("is_release_item"):
            release_name = str(task.get("release") or "Issues de Release").strip() or "Issues de Release"
            release_counts[release_name] += 1
            release_progress[release_name]["total"] += 1
            release_tasks += 1
            component = task.get("release_component") or classify_release_component(task)
            if component == "support":
                release_progress[release_name]["support_total"] += 1
            else:
                release_progress[release_name]["development_total"] += 1
            if bucket == "completed":
                release_completed += 1
                release_progress[release_name]["completed"] += 1
                if component == "support":
                    release_progress[release_name]["support_completed"] += 1
                else:
                    release_progress[release_name]["development_completed"] += 1
        else:
            non_release_tasks += 1
            if bucket == "completed":
                non_release_completed += 1
        if task.get("assignee") and task["assignee"] != "nan":
            assignee_counts[task["assignee"]] += 1

    completion_ratio = round((counters["completed"] / len(tasks)) * 100, 1) if tasks else 0
    release_completion_ratio = round((release_completed / release_tasks) * 100, 1) if release_tasks else 0
    non_release_completion_ratio = round((non_release_completed / non_release_tasks) * 100, 1) if non_release_tasks else 0
    release_weight = normalized_settings["release_weight"] / 100
    total_release_weight_points = sum(normalized_settings["release_weights"].get(name, 0) for name in release_names)
    weighted_release_completion_ratio = 0.0
    release_breakdown: list[dict[str, Any]] = []

    for release_name in release_names:
        progress = release_progress.get(release_name, {"total": 0, "completed": 0})
        total = int(progress["total"])
        completed = int(progress["completed"])
        ratio = round((completed / total) * 100, 1) if total else 0
        development_total = int(progress.get("development_total", 0))
        development_completed = int(progress.get("development_completed", 0))
        support_total = int(progress.get("support_total", 0))
        support_completed = int(progress.get("support_completed", 0))
        development_ratio = round((development_completed / development_total) * 100, 1) if development_total else 0
        support_ratio = round((support_completed / support_total) * 100, 1) if support_total else 0
        support_share = RELEASE_SUPPORT_CAP if support_total else 0.0
        development_share = 100.0 - support_share
        effective_ratio = round((development_ratio * (development_share / 100)) + (support_ratio * (support_share / 100)), 1)
        raw_weight = float(normalized_settings["release_weights"].get(release_name, 0))
        if total_release_weight_points > 0:
            normalized_weight = raw_weight / total_release_weight_points
        elif release_names:
            normalized_weight = 1 / len(release_names)
        else:
            normalized_weight = 0
        weighted_release_completion_ratio += effective_ratio * normalized_weight
        release_breakdown.append(
            {
                "name": release_name,
                "total": total,
                "completed": completed,
                "completion_ratio": ratio,
                "effective_completion_ratio": effective_ratio,
                "development_total": development_total,
                "development_completed": development_completed,
                "development_completion_ratio": development_ratio,
                "support_total": support_total,
                "support_completed": support_completed,
                "support_completion_ratio": support_ratio,
                "support_cap": RELEASE_SUPPORT_CAP,
                "weight": round(raw_weight, 2),
                "normalized_weight": round(normalized_weight * 100, 1),
            }
        )
    weighted_release_completion_ratio = round(weighted_release_completion_ratio, 1)

    if release_tasks and non_release_tasks:
        weighted_completion_ratio = round((weighted_release_completion_ratio * release_weight) + (non_release_completion_ratio * (1 - release_weight)), 1)
    elif release_tasks:
        weighted_completion_ratio = weighted_release_completion_ratio
    elif non_release_tasks:
        weighted_completion_ratio = non_release_completion_ratio
    else:
        weighted_completion_ratio = 0
    top_releases = release_counts.most_common(5)
    if not top_releases and releases:
        top_releases = sorted(((name, len(items)) for name, items in releases.items()), key=lambda item: item[1], reverse=True)[:5]

    return {
        "total": len(tasks),
        "completed": counters["completed"],
        "active": counters["active"],
        "overdue": counters["overdue"],
        "due_soon": counters["due_soon"],
        "upcoming": counters["upcoming"],
        "completion_ratio": completion_ratio,
        "weighted_completion_ratio": weighted_completion_ratio,
        "release_completion_ratio": release_completion_ratio,
        "weighted_release_completion_ratio": weighted_release_completion_ratio,
        "non_release_completion_ratio": non_release_completion_ratio,
        "release_weight": normalized_settings["release_weight"],
        "release_total": release_tasks,
        "non_release_total": non_release_tasks,
        "top_releases": top_releases,
        "release_breakdown": release_breakdown,
        "top_assignees": assignee_counts.most_common(5),
    }


def calculate_duration_days(start: str | None, end: str | None) -> int | None:
    start_date = parse_date(start)
    end_date = parse_date(end)
    if not start_date or not end_date:
        return None
    return (end_date - start_date).days + 1


def normalize_task_payload(task: dict[str, Any], default_line: int) -> dict[str, Any]:
    start_date = parse_date(task.get("start"))
    end_date = parse_date(task.get("end"))
    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date

    normalized = {
        "key": repair_text(task.get("key")),
        "title": repair_text(task.get("title")),
        "type": repair_text(task.get("type")),
        "parent": repair_text(task.get("parent")),
        "assignee": repair_text(task.get("assignee")),
        "status": normalize_status(task.get("status") or "Open"),
        "start": iso_date(start_date),
        "end": iso_date(end_date),
        "notes": repair_text(task.get("notes")),
        "duration_days": task.get("duration_days"),
        "release": repair_text(task.get("release")),
        "is_release_item": bool(task.get("is_release_item")),
        "release_component": repair_text(task.get("release_component")),
        "line": int(task.get("line") or default_line),
    }
    normalized["release"] = infer_release_name(normalized)
    normalized["is_release_item"] = bool(normalized["release"] or contains_release_keyword(normalized.get("title"), normalized.get("parent"), normalized.get("type")))
    normalized["release_component"] = classify_release_component(normalized)
    normalized["duration_days"] = calculate_duration_days(normalized["start"], normalized["end"])
    return normalized


def releases_from_tasks(tasks: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for task in tasks:
        release_name = (task.get("release") or "").strip()
        if release_name:
            grouped[release_name].append({"key": task.get("key", ""), "title": task.get("title", "")})
    return dict(grouped)


def release_lookup(releases: dict[str, list[dict[str, Any]]]) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for release_name, items in releases.items():
        for item in items:
            key = str(item.get("key") or "").strip()
            if key:
                lookup[key] = release_name
    return lookup


def merge_releases(existing_releases: dict[str, list[dict[str, Any]]], tasks: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    merged: dict[str, list[dict[str, Any]]] = defaultdict(list)
    task_keys = {str(task.get("key") or "").strip() for task in tasks}

    for release_name, items in (existing_releases or {}).items():
        for item in items:
            key = str(item.get("key") or "").strip()
            if key and key not in task_keys:
                merged[repair_text(release_name)].append({"key": repair_text(key), "title": repair_text(item.get("title"))})

    for task in tasks:
        release_name = repair_text(task.get("release"))
        if not release_name:
            continue
        merged[release_name].append({"key": repair_text(task.get("key")), "title": repair_text(task.get("title"))})

    return {release_name: items for release_name, items in merged.items() if items}


def finalize_project(
    project: str,
    source_name: str,
    source_type: str,
    settings: dict[str, Any],
    tasks: list[dict[str, Any]],
    releases: dict[str, list[dict[str, Any]]],
    *,
    imported_at: str | None = None,
    updated_at: str | None = None,
    slug: str | None = None,
) -> dict[str, Any]:
    today = date.today()
    for task in tasks:
        task["phase"] = derive_phase(task)
        task["bucket"] = classify_bucket(task, today)

    tasks.sort(key=lambda item: (int(item.get("line") or 999999), item.get("key") or "", item["title"]))
    normalized_settings = normalize_project_settings_for_releases(settings, release_names_from_sources(tasks, releases))
    metrics = compute_metrics(tasks, releases, normalized_settings)
    return NormalizedProject(
        project=project,
        source_name=repair_text(source_name),
        imported_at=imported_at or datetime.now().isoformat(timespec="seconds"),
        updated_at=updated_at or datetime.now().isoformat(timespec="seconds"),
        source_type=source_type,
        settings=normalized_settings,
        tasks=tasks,
        releases=releases,
        metrics=metrics,
        slug=slug,
    ).__dict__


def refresh_project(project_payload: dict[str, Any]) -> dict[str, Any]:
    existing_releases = project_payload.get("releases", {}) or {}
    existing_lookup = release_lookup(existing_releases)
    tasks = []
    for index, task in enumerate(project_payload.get("tasks", []), start=4):
        normalized = normalize_task_payload(task, index)
        if normalized["key"] and not normalized["release"]:
            normalized["release"] = existing_lookup.get(normalized["key"], "")
        if normalized["key"] and normalized["title"]:
            tasks.append(normalized)
    releases = merge_releases(existing_releases, tasks)
    return finalize_project(
        project=repair_text(project_payload.get("project", "Project")),
        source_name=project_payload.get("source_name", "manual"),
        source_type=project_payload.get("source_type", "manual"),
        settings=project_payload.get("settings"),
        tasks=tasks,
        releases=releases,
        imported_at=project_payload.get("imported_at"),
        updated_at=datetime.now().isoformat(timespec="seconds"),
        slug=project_payload.get("slug"),
    )


def parse_reference_workbook(file_bytes: bytes, source_name: str, project_name: str | None = None) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
    sheet = wb["Gantt (Weeks)"] if "Gantt (Weeks)" in wb.sheetnames else wb[wb.sheetnames[0]]
    releases: dict[str, list[dict[str, Any]]] = defaultdict(list)
    tasks: list[dict[str, Any]] = []

    release_maps: dict[str, set[str]] = {}
    for name in wb.sheetnames:
        if name.lower().startswith("release"):
            release_sheet = wb[name]
            keys = set()
            for row in release_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
                if row[0]:
                    keys.add(repair_text(row[0]))
                    releases[repair_text(name)].append({"key": repair_text(row[0]), "title": repair_text(row[1])})
            release_maps[name] = keys

    for idx, row in enumerate(sheet.iter_rows(min_row=4, max_col=10, values_only=True), start=4):
        key, title, item_type, parent, assignee, status, start, end, duration, notes = row
        if not key or not title:
            continue
        release_name = next((repair_text(name) for name, keys in release_maps.items() if repair_text(key) in keys), "")
        task = normalize_task_payload(
            {
            "key": repair_text(key),
            "title": repair_text(title),
            "type": repair_text(item_type),
            "parent": repair_text(parent),
            "assignee": repair_text(assignee),
            "status": normalize_status(status),
            "start": iso_date(parse_date(start)),
            "end": iso_date(parse_date(end)),
            "notes": repair_text(notes),
            "duration_days": int(duration) if duration not in (None, "") else None,
            "release": release_name,
            "line": idx,
            },
            idx,
        )
        tasks.append(task)
        if task.get("release") and not any(item["key"] == task["key"] for item in releases[task["release"]]):
            releases[task["release"]].append({"key": task["key"], "title": task["title"]})

    project = repair_text(project_name or Path(source_name).stem.replace("_", " "))
    return finalize_project(project, source_name, "excel", DEFAULT_PROJECT_SETTINGS, tasks, dict(releases))


def find_column(fieldnames: list[str], candidates: list[str]) -> str | None:
    lookup = {name.lower().strip(): name for name in fieldnames}
    for candidate in candidates:
        if candidate in lookup:
            return lookup[candidate]
    for name in fieldnames:
        lowered = name.lower().strip()
        if any(candidate in lowered for candidate in candidates):
            return name
    return None


def parse_jira_csv(file_bytes: bytes, source_name: str, project_name: str | None = None) -> dict[str, Any]:
    text = file_bytes.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []

    key_col = find_column(fieldnames, ["key", "issue key"])
    title_col = find_column(fieldnames, ["summary", "title"])
    type_col = find_column(fieldnames, ["issue type", "type"])
    parent_col = find_column(fieldnames, ["parent", "epic link", "epic name"])
    assignee_col = find_column(fieldnames, ["assignee"])
    status_col = find_column(fieldnames, ["status"])
    start_col = find_column(fieldnames, ["start date", "planned start", "start"])
    end_col = find_column(fieldnames, ["due date", "target end", "end date", "planned end", "finish"])
    duration_col = find_column(fieldnames, ["duration", "duration (d)", "story points"])
    release_col = find_column(fieldnames, ["fix version", "fix versions", "release", "target release"])
    notes_col = find_column(fieldnames, ["notes", "note", "comments", "comment", "justification", "justificacao"])

    tasks: list[dict[str, Any]] = []
    releases: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for idx, row in enumerate(reader, start=2):
        key = repair_text(row.get(key_col or "", ""))
        title = repair_text(row.get(title_col or "", ""))
        if not key or not title:
            continue
        release_name = repair_text(row.get(release_col or "", ""))
        duration = row.get(duration_col or "", "")
        task = normalize_task_payload(
            {
            "key": key,
            "title": title,
            "type": repair_text(row.get(type_col or "", "")),
            "parent": repair_text(row.get(parent_col or "", "")),
            "assignee": repair_text(row.get(assignee_col or "", "")),
            "status": normalize_status(row.get(status_col or "", "")),
            "start": iso_date(parse_date(row.get(start_col or "", ""))),
            "end": iso_date(parse_date(row.get(end_col or "", ""))),
            "notes": repair_text(row.get(notes_col or "", "")),
            "duration_days": int(float(duration)) if str(duration).strip() not in ("", "nan") and str(duration).replace(".", "", 1).isdigit() else None,
            "release": release_name,
            "line": idx,
            },
            idx,
        )
        tasks.append(task)
        if task.get("release"):
            releases[task["release"]].append({"key": key, "title": title})

    project = repair_text(project_name or Path(source_name).stem.replace("_", " "))
    return finalize_project(project, source_name, "csv", DEFAULT_PROJECT_SETTINGS, tasks, dict(releases))
